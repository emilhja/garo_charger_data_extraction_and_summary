"""Extract monthly invoice values from PDFs and write them to the summary workbook.

Expected invoice structure:
  2026/Fakturor/Övre/<MM - Mon>/<invoice>.pdf
  2026/Fakturor/Nedre/<MM - Mon>/<invoice>.pdf

Each month typically contains two PDFs: an Elhandel invoice and an Elnät invoice.
The script combines values from both documents into one monthly record per group.
"""

import re
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber

FAKTUROR_DIR = Path("2026/Fakturor")
SUMMARY_FILE_CANDIDATES = (
    Path("energy_2026_summary.xlsx"),
    Path("2026/energy_2026_summary.xlsx"),
    Path("energy_2026_summary_emil.xlsx"),
)
GROUP_KWH_COLUMN = {"Övre": 2, "Nedre": 3}
SHARED_RATE_COLUMNS = {
    "eloverforing_kr_per_kwh": 4,
    "elskatt_kr_per_kwh": 5,
    # The invoice does not expose Elcertifikat as a standalone line item.
    # We use `Rörliga kostnader*`, which explicitly includes Elcertifikat.
    "elcertifikat_kr_per_kwh": 6,
}
GROUP_ENERGY_PRICE_COLUMN = {
    "Övre": 7,
    "Nedre": 8,
}

MONTH_ROW = {
    1: 3,
    2: 4,
    3: 5,
    4: 6,
    5: 7,
    6: 8,
    7: 9,
    8: 10,
    9: 11,
    10: 12,
    11: 13,
    12: 14,
}

PER_KWH_LINE_RE = r"{label}\s+(\d{{4}})-(\d{{2}})-\d{{2}}\s+-\s+\d{{4}}-\d{{2}}-\d{{2}}\s+([\d\s]+)\s*kWh\s+([\d\s,]+)\s*öre/kWh\s+([\d\s,]+)"
FAST_AVGIFT_RE = r"Fast\s+Avgift\s+(\d{4})-(\d{2})-\d{2}\s+-\s+\d{4}-\d{2}-\d{2}\s+\d+\s+dagar\s+([\d\s,]+)\s*kr/år\s+([\d\s,]+)"


def parse_swedish_number(value: str) -> float:
    """Return a float from a Swedish-formatted numeric string."""
    return float(value.replace(" ", "").replace(",", "."))


def extract_text_from_pdf(pdf_path: Path) -> str:
    """Return the concatenated text from all pages in a PDF."""
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def extract_per_kwh_line(
    text: str,
    label: str,
) -> tuple[int, int, float, float] | None:
    """Return `(month, kwh, ore_per_kwh, amount_kr)` for a matching invoice row."""
    pattern = re.compile(
        PER_KWH_LINE_RE.format(label=label),
        re.IGNORECASE,
    )
    match = pattern.search(text)
    if not match:
        return None
    month = int(match.group(2))
    kwh = int(match.group(3).replace(" ", ""))
    ore_per_kwh = parse_swedish_number(match.group(4))
    amount_kr = parse_swedish_number(match.group(5))
    return month, kwh, ore_per_kwh, amount_kr


def extract_fast_avgift_line(text: str) -> tuple[int, float, float] | None:
    """Return `(month, kr_per_year, amount_kr)` for the Fast Avgift row."""
    match = re.search(FAST_AVGIFT_RE, text, re.IGNORECASE)
    if not match:
        return None
    month = int(match.group(2))
    kr_per_year = parse_swedish_number(match.group(3))
    amount_kr = parse_swedish_number(match.group(4))
    return month, kr_per_year, amount_kr


def extract_invoice_data_from_pdf(pdf_path: Path) -> tuple[int, dict[str, Any]] | None:
    """Return `(month, partial_data)` extracted from one invoice PDF, or `None`."""
    text = extract_text_from_pdf(pdf_path)
    normalized_text = text.upper()

    if "ELHANDEL" in normalized_text:
        spotpris = extract_per_kwh_line(text, "Spotpris")
        rorliga_kostnader = extract_per_kwh_line(text, r"Rörliga kostnader\*")
        fasta_paaslag = extract_per_kwh_line(text, "Fasta påslag")
        fast_avgift = extract_fast_avgift_line(text)

        if not spotpris:
            print(
                f"  Warning: {pdf_path.name} looks like Elhandel but no Spotpris row matched - check layout."
            )
            return None

        month, kwh, spotpris_ore, spotpris_amount = spotpris
        data = {
            "kwh": kwh,
            "spotpris_ore_per_kwh": spotpris_ore,
            "spotpris_amount_kr": spotpris_amount,
        }

        if rorliga_kostnader:
            row_month, row_kwh, ore_per_kwh, amount_kr = rorliga_kostnader
            if row_month != month or row_kwh != kwh:
                raise ValueError(
                    f"Unexpected `Rörliga kostnader*` month/kWh mismatch in {pdf_path}"
                )
            data["rorliga_kostnader_ore_per_kwh"] = ore_per_kwh
            data["rorliga_kostnader_amount_kr"] = amount_kr
            data["elcertifikat_kr_per_kwh"] = ore_per_kwh / 100

        if fasta_paaslag:
            row_month, row_kwh, ore_per_kwh, amount_kr = fasta_paaslag
            if row_month != month or row_kwh != kwh:
                raise ValueError(
                    f"Unexpected `Fasta påslag` month/kWh mismatch in {pdf_path}"
                )
            data["fasta_paaslag_ore_per_kwh"] = ore_per_kwh
            data["fasta_paaslag_amount_kr"] = amount_kr

        if fast_avgift:
            row_month, kr_per_year, amount_kr = fast_avgift
            if row_month != month:
                raise ValueError(
                    f"Unexpected `Fast avgift` month mismatch in {pdf_path}"
                )
            data["fast_avgift_elhandel_kr_per_year"] = kr_per_year
            data["fast_avgift_elhandel_amount_kr"] = amount_kr

        return month, data

    if "ELNÄT" in normalized_text or "ELNAT" in normalized_text:
        eloverforing = extract_per_kwh_line(text, "Elöverföring")
        if not eloverforing:
            print(
                f"  Warning: {pdf_path.name} looks like Elnät but no Elöverföring row matched - check layout."
            )
            return None

        month, kwh, eloverforing_ore, eloverforing_amount = eloverforing
        data = {
            "kwh": kwh,
            "eloverforing_kr_per_kwh": eloverforing_ore / 100,
            "eloverforing_amount_kr": eloverforing_amount,
        }

        elskatt = extract_per_kwh_line(text, r"(?:Elskatt|Energiskatt)")
        if elskatt:
            row_month, row_kwh, ore_per_kwh, amount_kr = elskatt
            if row_month != month or row_kwh != kwh:
                raise ValueError(
                    f"Unexpected `Elskatt/Energiskatt` month/kWh mismatch in {pdf_path}"
                )
            data["elskatt_kr_per_kwh"] = ore_per_kwh / 100
            data["elskatt_amount_kr"] = amount_kr

        fast_avgift = extract_fast_avgift_line(text)
        if fast_avgift:
            row_month, kr_per_year, amount_kr = fast_avgift
            if row_month != month:
                raise ValueError(
                    f"Unexpected `Fast avgift` month mismatch in {pdf_path}"
                )
            data["fast_avgift_elnat_kr_per_year"] = kr_per_year
            data["fast_avgift_elnat_amount_kr"] = amount_kr

        return month, data

    return None


def merge_invoice_month_data(
    existing_data: dict[str, Any] | None,
    new_data: dict[str, Any],
    pdf_path: Path,
) -> tuple[dict[str, Any], bool]:
    """Merge one invoice into one month record.

    Returns `(merged_data, is_duplicate_only)`.
    """
    if existing_data is None:
        return dict(new_data), False

    merged_data = dict(existing_data)
    duplicate_only = True
    for key, value in new_data.items():
        existing_value = merged_data.get(key)
        if existing_value is None:
            merged_data[key] = value
            duplicate_only = False
            continue
        if existing_value != value:
            raise ValueError(
                f"Conflicting invoice values for {key} in {pdf_path}: "
                f"{existing_value} vs {value}"
            )

    return merged_data, duplicate_only


def scan_fakturor() -> dict[str, dict[int, dict[str, Any]]]:
    """Scan invoice folders and return a nested mapping of group -> month -> invoice data."""
    usage_by_group_and_month: dict[str, dict[int, dict[str, Any]]] = {
        "Övre": {},
        "Nedre": {},
    }

    for group in ("Övre", "Nedre"):
        group_dir = FAKTUROR_DIR / group
        if not group_dir.exists():
            print(f"Warning: {group_dir} not found")
            continue

        for pdf_path in sorted(group_dir.rglob("*.pdf")):
            result = extract_invoice_data_from_pdf(pdf_path)
            if result:
                month, invoice_data = result
                existing_data = usage_by_group_and_month[group].get(month)
                merged_data, duplicate_only = merge_invoice_month_data(
                    existing_data, invoice_data, pdf_path
                )
                if duplicate_only:
                    print(
                        f"  Warning: duplicate invoice for {group} month={month} ignored ({pdf_path.name})"
                    )
                    continue
                usage_by_group_and_month[group][month] = merged_data
                print(
                    f"  {group} month={month}: {merged_data.get('kwh', 0)} kWh  ({pdf_path.name})"
                )

    return usage_by_group_and_month


def has_invoice_data(
    usage_by_group_and_month: dict[str, dict[int, dict[str, Any]]],
) -> bool:
    """Return `True` when at least one invoice total was extracted."""
    return any(usage_by_group_and_month[group] for group in usage_by_group_and_month)


def get_shared_rate_for_month(
    usage_by_group_and_month: dict[str, dict[int, dict[str, Any]]],
    month: int,
    field_name: str,
) -> float | None:
    """Return a shared rate for the month, ensuring groups do not disagree."""
    values = []
    for group in ("Övre", "Nedre"):
        value = usage_by_group_and_month[group].get(month, {}).get(field_name)
        if value is not None:
            values.append((group, value))

    if not values:
        return None

    first_group, first_value = values[0]
    for group, value in values[1:]:
        if value != first_value:
            raise ValueError(
                f"Conflicting {field_name} values in month {month}: "
                f"{first_group}={first_value}, {group}={value}"
            )
    return first_value


def resolve_summary_file() -> Path:
    """Return the first existing summary workbook path."""
    for candidate in SUMMARY_FILE_CANDIDATES:
        if candidate.exists():
            return candidate

    candidate_list = ", ".join(str(candidate) for candidate in SUMMARY_FILE_CANDIDATES)
    raise FileNotFoundError(
        f"Summary workbook not found. Checked: {candidate_list}."
    )


def update_summary_workbook(
    usage_by_group_and_month: dict[str, dict[int, dict[str, Any]]],
) -> None:
    """Write extracted invoice totals and rates into the `Grunddata` sheet."""
    summary_file = resolve_summary_file()

    workbook = openpyxl.load_workbook(summary_file)
    if "Grunddata" not in workbook.sheetnames:
        raise KeyError(
            f"`Grunddata` sheet not found in summary workbook: {summary_file}"
        )
    worksheet = workbook["Grunddata"]

    if usage_by_group_and_month["Övre"]:
        print(
            "\nÖvre invoice totals (informational only - Grunddata col B is left untouched):"
        )
        for month in sorted(usage_by_group_and_month["Övre"]):
            print(f"  Övre month {month}: {usage_by_group_and_month['Övre'][month]['kwh']} kWh")

    # Clear prior Nedre invoice imports before writing the current scan result.
    for row in MONTH_ROW.values():
        worksheet.cell(row=row, column=GROUP_KWH_COLUMN["Nedre"]).value = 0

    for group, month_values in usage_by_group_and_month.items():
        energy_price_column = GROUP_ENERGY_PRICE_COLUMN[group]
        for month, invoice_data in month_values.items():
            row = MONTH_ROW[month]
            if group == "Nedre":
                worksheet.cell(row=row, column=GROUP_KWH_COLUMN["Nedre"]).value = invoice_data["kwh"]

            spotpris = invoice_data.get("spotpris_ore_per_kwh")
            fasta_paaslag = invoice_data.get("fasta_paaslag_ore_per_kwh", 0.0)
            if spotpris is not None:
                worksheet.cell(row=row, column=energy_price_column).value = (
                    spotpris + fasta_paaslag
                ) / 100

    for month, row in MONTH_ROW.items():
        for field_name, column in SHARED_RATE_COLUMNS.items():
            value = get_shared_rate_for_month(usage_by_group_and_month, month, field_name)
            if value is not None:
                worksheet.cell(row=row, column=column).value = value

    workbook.save(summary_file)
    print(f"\nSaved {summary_file}")


def main() -> None:
    """Run the invoice extraction flow and update the workbook if data is found."""
    print("Scanning invoices...\n")
    usage_by_group_and_month = scan_fakturor()

    if not has_invoice_data(usage_by_group_and_month):
        print("No Elhandel invoices found.")
        return

    update_summary_workbook(usage_by_group_and_month)
    print("\nGrunddata invoice values written:")
    for group in ("Övre", "Nedre"):
        for month in sorted(usage_by_group_and_month[group]):
            invoice_data = usage_by_group_and_month[group][month]
            details = [
                f"kWh={invoice_data['kwh']}",
            ]
            for field_name in (
                "eloverforing_kr_per_kwh",
                "elskatt_kr_per_kwh",
                "elcertifikat_kr_per_kwh",
                "spotpris_ore_per_kwh",
                "rorliga_kostnader_ore_per_kwh",
                "fasta_paaslag_ore_per_kwh",
                "fast_avgift_elhandel_amount_kr",
                "fast_avgift_elnat_amount_kr",
            ):
                value = invoice_data.get(field_name)
                if value is not None:
                    details.append(f"{field_name}={value}")
            print(f"  {group} month {month}: " + ", ".join(details))


if __name__ == "__main__":
    main()
