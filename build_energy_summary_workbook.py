"""Create a yearly Excel summary workbook from charger usage and price inputs."""

from collections import defaultdict
import json
from pathlib import Path
import re
from typing import Any

import openpyxl
from check_anomalies import run_checks, print_report, STANDING_ISSUES

YEAR = 2026
CONFIG_FILE = Path("garage_config.json")
DATA_DIR = Path("data")
ENERGY_FILE = DATA_DIR / f"energy_{YEAR}.json"
PRICES_FILE = Path(f"{YEAR}/Förbrukning/Priser och Förbrukningsuppgifter laddstationer {YEAR}.xlsx")
SUMMARY_OUTPUT_FILE = Path(f"energy_{YEAR}_summary.xlsx")
MONTH_ABBREVIATIONS = ["Jan", "Feb", "Mars", "April", "Maj", "Juni", "Juli", "Aug", "Sep", "Okt", "Nov", "Dec"]
MONTH_NAMES = ["Jan", "Februari", "Mars", "April", "Maj", "Juni", "Juli", "Augusti", "September", "Oktober", "November", "December"]


def load_garage_configuration() -> dict[str, dict[str, dict[str, Any]]]:
    """Return the full garage configuration grouped by `övre` and `nedre`."""
    with CONFIG_FILE.open(encoding="utf-8") as config_file:
        return json.load(config_file)


def build_garage_metadata(
    garage_config: dict[str, dict[str, dict[str, Any]]]
) -> dict[str, tuple[str, str]]:
    """Build a garage -> (mgg, group) lookup from the configuration file."""
    garage_metadata = {}
    for garage_group in ("övre", "nedre"):
        for garage_name, garage_info in garage_config.get(garage_group, {}).items():
            garage_address = garage_info.get("mgg") or garage_info.get("adress")
            if not garage_address:
                raise KeyError(
                    f"Missing address for {garage_name}. Expected `mgg` or `adress` in garage_config.json."
                )
            garage_metadata[garage_name] = (garage_address, garage_group)
    return garage_metadata


def garage_sort_key(garage_name: str) -> tuple[int, int | float, str]:
    """Return a sort key that keeps `Garage02` before `Garage10`."""
    match = re.search(r"(\d+)$", garage_name)
    if match:
        return (0, int(match.group(1)), garage_name)
    return (1, float("inf"), garage_name)


def make_garage_sheet_name(garage_name: str, garage_address: str) -> str:
    """Return the normalized Excel sheet title for a garage."""
    garage_number = garage_name.replace("Garage", "").zfill(2)
    return f"{garage_address} - garage {garage_number}"


def load_monthly_prices() -> tuple[dict[int, float], dict[int, float]]:
    """Read monthly cost per kWh excluding VAT from the source price workbook."""
    workbook = openpyxl.load_workbook(PRICES_FILE, data_only=True)
    if "Grunddata" not in workbook.sheetnames:
        raise KeyError(f"`Grunddata` sheet not found in pricing workbook: {PRICES_FILE}")
    worksheet = workbook["Grunddata"]
    # Rows 3-14 map to January-December.
    # Column I stores the upper garage price and column K stores the lower garage price.
    upper_group_prices = {}
    lower_group_prices = {}
    for row_idx, month_num in enumerate(range(1, 13), start=3):
        upper_price = worksheet.cell(row=row_idx, column=9).value
        lower_price = worksheet.cell(row=row_idx, column=11).value
        upper_group_prices[month_num] = upper_price or 0
        lower_group_prices[month_num] = lower_price or 0
    workbook.close()
    return upper_group_prices, lower_group_prices


def load_monthly_energy_usage() -> dict[str, dict[int, float]]:
    """Aggregate the yearly JSON export into monthly kWh totals per garage."""
    with ENERGY_FILE.open(encoding="utf-8") as energy_file:
        energy_rows = json.load(energy_file)
    monthly_usage_by_garage = defaultdict(lambda: defaultdict(float))
    for energy_row in energy_rows:
        month_num = int(energy_row["date"][5:7])
        monthly_usage_by_garage[energy_row["name"]][month_num] += energy_row.get("energy_kwh") or 0
    return monthly_usage_by_garage


def create_summary_workbook(
    monthly_usage_by_garage: dict[str, dict[int, float]],
    output_file: Path = SUMMARY_OUTPUT_FILE,
) -> None:
    """Create the summary workbook with one sheet per garage plus aggregate sheets.

    Row 3 of each garage sheet uses Excel formulas that multiply row 2 kWh by the
    matching monthly price in the `Grunddata` sheet (column I for övre, K for nedre).
    This keeps costs live in Excel: editing a kWh value or a price automatically
    recalculates all cost cells without regenerating the workbook.
    """
    garage_config = load_garage_configuration()
    garage_metadata = build_garage_metadata(garage_config)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # Include all configured garages even if the energy export has no readings yet.
    all_garage_names = set(monthly_usage_by_garage.keys()) | set(garage_metadata.keys())

    for garage_name in sorted(all_garage_names, key=garage_sort_key):
        mgg_name, garage_group = garage_metadata.get(garage_name, ("Unknown", "nedre"))
        # Column I = övre price per kWh; column K = nedre price per kWh (Grunddata rows 3–14).
        price_col = "I" if garage_group == "övre" else "K"
        sheet_name = make_garage_sheet_name(garage_name, mgg_name)

        worksheet = workbook.create_sheet(title=sheet_name)

        worksheet["A1"] = "Månad"
        for column_index, month_label in enumerate(MONTH_ABBREVIATIONS, start=2):
            worksheet.cell(row=1, column=column_index, value=month_label)

        worksheet["A2"] = "Förbrukning"
        for column_index, month_num in enumerate(range(1, 13), start=2):
            monthly_kwh = monthly_usage_by_garage.get(garage_name, {}).get(month_num)
            usage_cell = worksheet.cell(row=2, column=column_index, value=monthly_kwh if monthly_kwh is not None else 0)
            usage_cell.number_format = "0.0"

        worksheet["A3"] = "Kostnad exkl. moms (se Blad 1)"
        for column_index, month_num in enumerate(range(1, 13), start=2):
            kwh_col = openpyxl.utils.get_column_letter(column_index)
            # Grunddata rows 3–14 correspond to months 1–12.
            grunddata_row = month_num + 2
            formula = f"={kwh_col}2*Grunddata!{price_col}{grunddata_row}"
            worksheet.cell(row=3, column=column_index, value=formula)

        worksheet.column_dimensions["A"].width = 34

    add_cost_summary_sheet(workbook, garage_config, garage_metadata)
    add_grunddata_sheet(workbook, garage_config)

    workbook.save(output_file)
    print(f"Saved {output_file} with {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")


def add_cost_summary_sheet(
    workbook: openpyxl.Workbook,
    garage_config: dict[str, dict[str, dict[str, Any]]],
    garage_metadata: dict[str, tuple[str, str]],
) -> None:
    """Add the front summary sheet that references monthly cost rows from each garage sheet."""
    worksheet = workbook.create_sheet(title="Samlad bild i SEK", index=0)

    worksheet["A1"] = "Adress & Garage / Månad"
    for column_index, month_label in enumerate(MONTH_NAMES, start=2):
        worksheet.cell(row=1, column=column_index, value=month_label)

    apartment_number_by_garage = {}
    for garage_group in ("övre", "nedre"):
        for garage_name, garage_info in garage_config.get(garage_group, {}).items():
            apartment_number_by_garage[garage_name] = garage_info["lgh"]

    sorted_garages = sorted(
        garage_metadata.items(),
        key=lambda garage_item: apartment_number_by_garage.get(garage_item[0], 999),
    )

    for row_index, (garage_name, (mgg_name, _garage_group)) in enumerate(sorted_garages, start=2):
        apartment_number = apartment_number_by_garage.get(garage_name, "?")
        sheet_name = make_garage_sheet_name(garage_name, mgg_name)
        garage_number = garage_name.replace("Garage", "").zfill(2)
        address_label = f"{apartment_number} {mgg_name} - garage {garage_number}"

        worksheet.cell(row=row_index, column=1, value=address_label)

        # Each summary cell links directly to row 3 in the garage sheet so Excel recalculates
        # automatically if any monthly values or price inputs change.
        for column_offset, _month in enumerate(MONTH_ABBREVIATIONS):
            garage_column_letter = openpyxl.utils.get_column_letter(column_offset + 2)
            formula = f"='{sheet_name}'!{garage_column_letter}3"
            worksheet.cell(row=row_index, column=column_offset + 2, value=formula)

    worksheet.column_dimensions["A"].width = 34


def add_grunddata_sheet(
    workbook: openpyxl.Workbook,
    garage_config: dict[str, dict[str, dict[str, Any]]],
) -> None:
    """Copy the pricing sheet structure and replace kWh columns with live SUM formulas."""
    source_workbook = openpyxl.load_workbook(PRICES_FILE, data_only=True)
    if "Grunddata" not in source_workbook.sheetnames:
        raise KeyError(f"`Grunddata` sheet not found in pricing workbook: {PRICES_FILE}")
    source_worksheet = source_workbook["Grunddata"]

    worksheet = workbook.create_sheet(title="Grunddata", index=1)

    upper_group_sheet_names = []
    lower_group_sheet_names = []
    for garage_name, garage_info in garage_config.get("övre", {}).items():
        garage_address = garage_info.get("mgg") or garage_info.get("adress")
        if not garage_address:
            raise KeyError(f"Missing address for {garage_name} in garage_config.json.")
        upper_group_sheet_names.append(make_garage_sheet_name(garage_name, garage_address))
    for garage_name, garage_info in garage_config.get("nedre", {}).items():
        garage_address = garage_info.get("mgg") or garage_info.get("adress")
        if not garage_address:
            raise KeyError(f"Missing address for {garage_name} in garage_config.json.")
        lower_group_sheet_names.append(make_garage_sheet_name(garage_name, garage_address))

    if not upper_group_sheet_names:
        raise ValueError("No `övre` garages found in garage_config.json.")
    if not lower_group_sheet_names:
        raise ValueError("No `nedre` garages found in garage_config.json.")

    for column_index in range(1, 13):
        worksheet.cell(row=1, column=column_index, value=source_worksheet.cell(row=1, column=column_index).value)

    for column_index in range(1, 13):
        worksheet.cell(row=2, column=column_index, value=source_worksheet.cell(row=2, column=column_index).value)

    for source_row in range(3, 15):
        month_index = source_row - 3
        month_col_letter = openpyxl.utils.get_column_letter(month_index + 2)

        worksheet.cell(row=source_row, column=1, value=source_worksheet.cell(row=source_row, column=1).value)

        # Col B: sum kWh row from all övre garage sheets (GARO API data).
        upper_group_references = "+".join(f"'{sheet_name}'!{month_col_letter}2" for sheet_name in upper_group_sheet_names)
        worksheet.cell(row=source_row, column=2, value=f"={upper_group_references}")

        # Col C: nedre kWh comes from invoice PDFs, not from the garage sheets.
        # Leave as 0 here; extract_invoice_kwh.py writes the real values.
        worksheet.cell(row=source_row, column=3, value=0)

        for column_index in range(4, 13):
            worksheet.cell(row=source_row, column=column_index, value=source_worksheet.cell(row=source_row, column=column_index).value)

    for source_row in range(15, 20):
        for column_index in range(1, 13):
            cell_value = source_worksheet.cell(row=source_row, column=column_index).value
            if cell_value is not None:
                worksheet.cell(row=source_row, column=column_index, value=cell_value)

    worksheet.column_dimensions["A"].width = 20
    source_workbook.close()


def main() -> None:
    """Run the workbook generation flow from current JSON and pricing inputs."""
    if not ENERGY_FILE.exists():
        raise FileNotFoundError(f"Energy data not found: {ENERGY_FILE}. Run fetch_garo_energy.py first.")
    if not PRICES_FILE.exists():
        raise FileNotFoundError(f"Pricing workbook not found: {PRICES_FILE}")
    upper_group_prices, lower_group_prices = load_monthly_prices()
    print("Övre prices (cost/kWh excl. VAT):")
    for month_num, price in upper_group_prices.items():
        print(f"  {MONTH_ABBREVIATIONS[month_num - 1]:10s}: {price}")
    print("Nedre prices (cost/kWh excl. VAT):")
    for month_num, price in lower_group_prices.items():
        print(f"  {MONTH_ABBREVIATIONS[month_num - 1]:10s}: {price}")

    monthly_usage_by_garage = load_monthly_energy_usage()
    create_summary_workbook(monthly_usage_by_garage)

    anomalies = run_checks()
    print_report(anomalies)


if __name__ == '__main__':
    main()
